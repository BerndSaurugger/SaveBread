from pandas import read_excel, read_csv
import openpyxl
import os


def file_path():
    """
    Automatically generates the filepath of the saved file
    """
    list_of_data = os.listdir("src/data")
    count_input_files = 0
    for i in range(len(list_of_data)):
        if list_of_data[i].endswith(('.xlsx','.xls', '.csv')) and list_of_data[i] != "savebread_output.xlsx":
            file_name = list_of_data[i]
            count_input_files += 1
    assert count_input_files != 0, "please provide a file"
    assert count_input_files == 1, "two many files provided"
    file_path = "src/data/" + str(file_name)
    return file_path


def load_alternative_file(file_path):
    """
    Please provide a dataset from type xlsx, xls or csv and make sure the first row name your colomns
    """
    # assert file type and load data 
    # assert file.endswith(('.xlsx','.xls', '.csv')), "Filetype must be xlsx or csv"
    if file_path.endswith('.csv'):
        df_file = read_csv(file_path)
    else:
        df_file = read_excel(file_path, sheet_name=0)


    # assert correct reading of the dataset
    counter = 0
    for i in range(len(df_file.columns)):
        if df_file.columns[i] == "Unnamed: {}".format(i):
            counter += 1
    assert counter != len(df_file.columns), "Fileread Error: Please check your data! Your Column names (first row in the sheet) must not be empty"

    # assert all rows/cols are imported
    if file_path.endswith(('.xlsx','.xls')):
        assert len(df_file.columns) == openpyxl.load_workbook(file_path).get_sheet_by_name(all_worksheets[0]).max_column(), "Error while reading the columns"
        assert len(df_file) != openpyxl.load_workbook(file_path).get_sheet_by_name(all_worksheets[0]).max_rows, "Error while reading the rows"

    # assert empty dataset
    assert len(df_file) != 0, "Error: Empty dataset inserted! Please check your data or ask the next Data Scientist"

    # assert missing values
    assert df_file.isna().any().any() == False, "The Dataset has missing values" # I am not sure if we should not go for an imputing strategy but anyway: accordingto the issue!

    return df_file


def load_file():
    """
    Loads the data by using the alternatively generated filepath
    """
    return load_alternative_file(file_path())
