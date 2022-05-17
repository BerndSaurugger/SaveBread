import unittest
import openpyxl
import xlrd as xl

from filehandler import load_input


# get_file_path()

# load_alternative_file("src/feature_list.xlsx")
# test_df = load_alternative_file("src/feature_list.xlsx")
# print(alternative_file_path)
# col= openpyxl.load_workbook(alternative_file_path).worksheets[0].max_column
# row= openpyxl.load_workbook(file_path).worksheets[0].max_column

class TestInput(unittest.TestCase):
    # def test_input_amount(self):
    #     if load_input.count_input_files() == 0:
    #         raise FileNotFoundError("File not correctly uploaded!")
    #     if load_input.count_input_files() > 1: 
    #         raise ValueError("You have provided {} files, but there is only one alowed".format(load_input.count_input_files()))

    def test_returns_0_on_emty_folder(self):
        assert load_input.count_input_files("src/test_data/folder_with_nan") == 0
        
    def test_raises_exeption_on_the_empty_folder(self):
        with self.assertRaises(FileNotFoundError) as context: 
            load_input.get_file_path("src/test_data/folder_with_nan")

    def test_raises_value_on_to_many_files(self):
        with self.assertRaises(ValueError) as context: 
            load_input.get_file_path("src/test_data/folder_with_1_xlsx_and_1_csv_file")

    def test_returns_1_on_folder_with_1_file(self):
        assert load_input.count_input_files("src/test_data/folder_with_1_xlsx_file") == 1

    def test_first_name_of_file_empty(self):
        data_frame = load_input.load_file("src/test_data/folder_with_files/feature_list_blanc_fist_row.xlsx")
        counter = 0
        for i in range(len(data_frame.columns)):
            if data_frame.columns[i] == "Unnamed: {}".format(i):
                counter += 1 
        assert counter == len(data_frame.columns)

    def test_first_name_of_file_not_empty(self):
        data_frame = load_input.load_file("src/test_data/folder_with_files/feature_list_all_correct.xlsx")
        counter = 0
        for i in range(len(data_frame.columns)):
            if data_frame.columns[i] == "Unnamed: {}".format(i):
                counter += 1 
        assert counter != len(data_frame.columns) 

    def test_empty_dataset(self):
        data_frame = load_input.load_file("src/test_data/folder_with_files/empty_dataset.xlsx")
        assert len(data_frame) == 0

    def test_filled_dataset(self):
        data_frame = load_input.load_file("src/test_data/folder_with_files/feature_list_all_correct.xlsx")
        assert len(data_frame) != 0

    def test_missing_values(self):
        data_frame = load_input.load_file("src/test_data/folder_with_files/feature_list_all_nan.xlsx")
        assert data_frame.isna().any().any() == True

    def test_no_missing_values(self):
        data_frame = load_input.load_file("src/test_data/folder_with_files/feature_list_all_correct.xlsx")
        assert data_frame.isna().any().any() == False

    def test_all_columns_correctly_read(self):
        data_frame = load_input.load_file("src/test_data/folder_with_files/feature_list_all_correct.xlsx")
        wb = xl.open_workbook("src/test_data/folder_with_files/feature_list_all_correct.xlsx").sheet_by_index(0)
        wb.cell_value(0,0)
        assert len(data_frame.columns) == wb.ncols, "Error while reading the columns"

    def test_all_rows_correctly_read(self):
        data_frame = load_input.load_file("src/test_data/folder_with_files/feature_list_all_correct.xlsx")
        wb = xl.open_workbook("src/test_data/folder_with_files/feature_list_all_correct.xlsx").sheet_by_index(0)
        wb.cell_value(0,0)
        assert len(data_frame) == wb.nrows-1, "Error while reading the rows" #care: pd.DataFrame uses 1 Row as Featurename


if __name__ == '__main__':
    unittest.main()